# ui/report_window.py — PDF / Excel 리포트 생성 창
from __future__ import annotations
import os
import numpy as np
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QGroupBox, QFormLayout, QCheckBox,
    QFileDialog, QProgressBar, QMessageBox,
)
from PyQt5.QtCore import Qt, QThread, QObject, pyqtSignal
from ui.dialogs import DARK, PANEL, TEXT, MUTED, BORDER, STYLE_DLG

from core.utils import SF_SENS



def _pr_to_sf(pr: float) -> str:
    if pr <= -999:
        return "미커버"
    for sf in sorted(SF_SENS.keys()):
        if pr >= SF_SENS[sf]:
            return f"SF{sf}"
    return "불가"


# ── 리포트 워커 ──────────────────────────────────────────────
class ReportWorker(QObject):
    sig_progress = pyqtSignal(int, str)
    sig_done     = pyqtSignal(str)
    sig_err      = pyqtSignal(str)

    def __init__(self, mode, path, result, gws, nodes, heatmaps, settings):
        super().__init__()
        self.mode      = mode       # 'pdf' | 'excel'
        self.path      = path
        self.result    = result
        self.gws       = gws
        self.nodes     = nodes
        self.heatmaps  = heatmaps
        self.settings  = settings

    def run(self):
        try:
            if self.mode == 'excel':
                self._export_excel()
            else:
                self._export_pdf()
            self.sig_done.emit(self.path)
        except Exception:
            import traceback
            self.sig_err.emit(traceback.format_exc())

    # ── Excel 내보내기 ───────────────────────────────────────
    def _export_excel(self):
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── 색상 정의 ────────────────────────────────────────
        HDR_FILL  = PatternFill("solid", fgColor="1E2130")
        COV_FILL  = PatternFill("solid", fgColor="0D2010")
        UCOV_FILL = PatternFill("solid", fgColor="200D0D")
        HDR_FONT  = Font(color="7AB8E8", bold=True, size=10)
        COV_FONT  = Font(color="00C94A", size=10)
        UCOV_FONT = Font(color="FF4444", size=10)
        CTR_ALIGN = Alignment(horizontal="center", vertical="center")

        def _thin_border():
            s = Side(style="thin", color="2A2F3B")
            return Border(left=s, right=s, top=s, bottom=s)

        def _hdr(ws, row, cols):
            for c, v in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill      = HDR_FILL
                cell.font      = HDR_FONT
                cell.alignment = CTR_ALIGN
                cell.border    = _thin_border()

        # ─── 시트 1: 요약 ────────────────────────────────────
        self.sig_progress.emit(10, "요약 시트 작성 중...")
        ws1 = wb.active
        ws1.title = "커버리지 요약"

        result  = self.result
        n_total = result.n_total
        n_cov   = result.n_covered
        pct     = result.coverage_pct
        nodes_r = result.nodes

        n_multi    = sum(1 for nd in nodes_r if nd.n_rx_gw >= 2)
        n_single   = sum(1 for nd in nodes_r if nd.covered and nd.n_rx_gw == 1)
        n_uncov    = n_total - n_cov
        ovlp_pct   = (n_multi / n_cov * 100) if n_cov > 0 else 0.0
        single_pct = (n_single / n_total * 100) if n_total > 0 else 0.0

        summary = [
            ("항목", "값", "단위"),
            ("전체 Node 수",      n_total,           "개"),
            ("커버 Node 수",      n_cov,             "개"),
            ("전체 커버리지",     f"{pct:.1f}",      "%"),
            ("음영 Node 수",      n_uncov,           "개"),
            ("중첩 커버 Node",    n_multi,           "개"),
            ("중첩 커버 비율",    f"{ovlp_pct:.1f}", "%"),
            ("단독 커버 비율",    f"{single_pct:.1f}","%"),
            ("활성 GW 수",        len(self.gws),     "개"),
            ("평균 수신 GW 수",   f"{getattr(result,'avg_n_rx_gw',0):.2f}", "개"),
            ("매크로 다이버시티", f"{getattr(result,'macro_diversity_gain',0):.2f}", "dB"),
            ("평균 ToA",          f"{getattr(result,'avg_toa_ms',0):.1f}", "ms"),
        ]

        _hdr(ws1, 1, ["항목", "값", "단위"])
        for r, row_data in enumerate(summary[1:], 2):
            for c, val in enumerate(row_data, 1):
                cell = ws1.cell(row=r, column=c, value=val)
                cell.alignment = CTR_ALIGN
                cell.border    = _thin_border()

        ws1.column_dimensions['A'].width = 22
        ws1.column_dimensions['B'].width = 14
        ws1.column_dimensions['C'].width = 8

        # ─── 시트 2: GW 목록 ─────────────────────────────────
        self.sig_progress.emit(25, "GW 시트 작성 중...")
        ws2 = wb.create_sheet("GW 목록")
        gw_cols = ["Callsign", "경도", "위도", "Pt(dBm)", "Gt(dBi)",
                   "Lt(dB)", "높이(m)", "담당 Node 수"]
        _hdr(ws2, 1, gw_cols)
        gw_counts = result.gw_counts
        for r, gw in enumerate(self.gws, 2):
            cnt = gw_counts.get(gw.callsign, 0)
            row_vals = [gw.callsign, round(gw.lon, 6), round(gw.lat, 6),
                        gw.pt_dbm, gw.gt_dbi, gw.lt_db, gw.hb_m, cnt]
            for c, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c, value=val)
                cell.alignment = CTR_ALIGN
                cell.border    = _thin_border()
        for i, w in enumerate([14,14,14,10,10,10,10,12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # ─── 시트 3: Node 목록 ───────────────────────────────
        self.sig_progress.emit(50, "Node 시트 작성 중...")
        ws3 = wb.create_sheet("Node 목록")
        nd_cols = ["Callsign", "경도", "위도", "Gr(dBi)", "Lr(dB)",
                   "높이(m)", "최소수신(dBm)", "연결 GW", "수신전력(dBm)", "SF 등급", "상태"]
        _hdr(ws3, 1, nd_cols)
        for r, (nd, info) in enumerate(zip(self.nodes, nodes_r), 2):
            best_gw  = info.best_gw or "─"
            best_pr  = round(info.best_pr, 1) if info.best_pr > -999 else "─"
            sf_grade = _pr_to_sf(info.best_pr) if info.best_pr > -999 else "─"
            status   = "커버" if info.covered else "미커버"
            row_vals = [
                nd.callsign, round(nd.lon, 6), round(nd.lat, 6),
                nd.gr_dbi, nd.lr_db, nd.hm_m, nd.min_rx_dbm,
                best_gw, best_pr, sf_grade, status
            ]
            for c, val in enumerate(row_vals, 1):
                cell = ws3.cell(row=r, column=c, value=val)
                cell.alignment = CTR_ALIGN
                cell.border    = _thin_border()
                if c == 11:
                    if info.covered:
                        cell.fill = COV_FILL
                        cell.font = COV_FONT
                    else:
                        cell.fill = UCOV_FILL
                        cell.font = UCOV_FONT
        for i, w in enumerate([14,14,14,10,10,10,14,14,14,10,10], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        # ─── 시트 4: SF별 분포 ───────────────────────────────
        self.sig_progress.emit(75, "SF 분포 시트 작성 중...")
        ws4 = wb.create_sheet("SF 분포")
        sf_cols = ["SF", "감도(dBm)", "커버 Node 수", "비율(%)"]
        _hdr(ws4, 1, sf_cols)
        for r, (sf, sens) in enumerate(SF_SENS.items(), 2):
            n_sf   = sum(1 for nd in nodes_r if nd.best_pr >= sens)
            pct_sf = n_sf / n_total * 100 if n_total else 0
            for c, val in enumerate(
                    [f"SF{sf}", sens, n_sf, f"{pct_sf:.1f}"], 1):
                cell = ws4.cell(row=r, column=c, value=val)
                cell.alignment = CTR_ALIGN
                cell.border    = _thin_border()
        for i, w in enumerate([8, 12, 14, 10], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w

        self.sig_progress.emit(90, "파일 저장 중...")
        wb.save(self.path)
        self.sig_progress.emit(100, "완료")

    # ── PDF 내보내기 ─────────────────────────────────────────
    def _export_pdf(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import platform

        # 한글 폰트 등록
        font_name = "Helvetica"
        if platform.system() == "Windows":
            font_path = "C:/Windows/Fonts/malgun.ttf"
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont("MalgunGothic", font_path))
                    font_name = "MalgunGothic"
                except Exception:
                    pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title2", parent=styles["Title"],
            fontName=font_name, fontSize=16, spaceAfter=6)
        h2_style = ParagraphStyle(
            "H2", parent=styles["Heading2"],
            fontName=font_name, fontSize=12, spaceAfter=4)
        body_style = ParagraphStyle(
            "Body", parent=styles["Normal"],
            fontName=font_name, fontSize=9, spaceAfter=2)

        doc   = SimpleDocTemplate(self.path, pagesize=A4,
                                  leftMargin=15*mm, rightMargin=15*mm,
                                  topMargin=15*mm, bottomMargin=15*mm)
        story = []

        # ── 제목 ─────────────────────────────────────────────
        story.append(Paragraph("LoRaScape 커버리지 분석 리포트", title_style))
        story.append(Paragraph("SmartCity LoRaWAN Network Simulator — SOLUWINS", body_style))
        story.append(Spacer(1, 8*mm))

        self.sig_progress.emit(20, "PDF 요약 섹션 작성 중...")

        result  = self.result
        n_total = result.n_total
        n_cov   = result.n_covered
        pct     = result.coverage_pct
        nodes_r = result.nodes

        n_multi  = sum(1 for nd in nodes_r if nd.n_rx_gw >= 2)
        n_uncov  = n_total - n_cov
        ovlp_pct = (n_multi / n_cov * 100) if n_cov > 0 else 0.0

        # ── 요약 테이블 ──────────────────────────────────────
        story.append(Paragraph("1. 커버리지 요약", h2_style))
        summary_data = [
            ["항목", "값", "단위"],
            ["전체 Node 수",     str(n_total),          "개"],
            ["커버 Node 수",     str(n_cov),            "개"],
            ["전체 커버리지",    f"{pct:.1f}",          "%"],
            ["음영 Node 수",     str(n_uncov),          "개"],
            ["중첩 커버 Node",   str(n_multi),          "개"],
            ["중첩 커버 비율",   f"{ovlp_pct:.1f}",     "%"],
            ["활성 GW 수",       str(len(self.gws)),    "개"],
            ["평균 ToA",         f"{getattr(result,'avg_toa_ms',0):.1f}", "ms"],
        ]
        tbl_s = Table(summary_data, colWidths=[80*mm, 50*mm, 30*mm])
        tbl_s.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E2130')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.HexColor('#7AB8E8')),
            ('FONTNAME',   (0,0), (-1,-1), font_name),
            ('FONTSIZE',   (0,0), (-1,-1), 9),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#2A2F3B')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.HexColor('#181B22'), colors.HexColor('#1E2130')]),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#E0E4EF')),
        ]))
        story.append(tbl_s)
        story.append(Spacer(1, 6*mm))

        # ── GW 목록 ──────────────────────────────────────────
        self.sig_progress.emit(45, "PDF GW 섹션 작성 중...")
        story.append(Paragraph("2. GW 목록", h2_style))
        gw_data = [["Callsign", "경도", "위도", "Pt(dBm)", "높이(m)", "담당 Node"]]
        for gw in self.gws:
            cnt = result.gw_counts.get(gw.callsign, 0)
            gw_data.append([
                gw.callsign,
                f"{gw.lon:.5f}", f"{gw.lat:.5f}",
                str(gw.pt_dbm), f"{gw.hb_m:.1f}m", f"{cnt}개"
            ])
        tbl_gw = Table(gw_data,
                       colWidths=[35*mm,35*mm,35*mm,25*mm,25*mm,25*mm])
        tbl_gw.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E2130')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.HexColor('#7AB8E8')),
            ('FONTNAME',   (0,0), (-1,-1), font_name),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#2A2F3B')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.HexColor('#181B22'), colors.HexColor('#1E2130')]),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#E0E4EF')),
        ]))
        story.append(tbl_gw)
        story.append(Spacer(1, 6*mm))

        # ── Node 목록 ────────────────────────────────────────
        self.sig_progress.emit(70, "PDF Node 섹션 작성 중...")
        story.append(Paragraph("3. Node 커버리지 목록 (상위 50개)", h2_style))
        nd_data = [["Callsign", "연결 GW", "수신전력(dBm)", "SF 등급", "상태"]]
        for nd, info in zip(self.nodes[:50], nodes_r[:50]):
            best_pr  = f"{info.best_pr:.1f}" if info.best_pr > -999 else "─"
            sf_grade = _pr_to_sf(info.best_pr) if info.best_pr > -999 else "─"
            status   = "커버" if info.covered else "미커버"
            nd_data.append([
                nd.callsign, info.best_gw or "─",
                best_pr, sf_grade, status
            ])
        tbl_nd = Table(nd_data,
                       colWidths=[35*mm,40*mm,35*mm,25*mm,25*mm])
        tbl_nd.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E2130')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.HexColor('#7AB8E8')),
            ('FONTNAME',   (0,0), (-1,-1), font_name),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#2A2F3B')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.HexColor('#181B22'), colors.HexColor('#1E2130')]),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#E0E4EF')),
        ]))
        story.append(tbl_nd)

        self.sig_progress.emit(90, "PDF 저장 중...")
        doc.build(story)
        self.sig_progress.emit(100, "완료")


# ── 리포트 창 ────────────────────────────────────────────────
class ReportWindow(QDialog):
    def __init__(self, result, gws, nodes, heatmaps=None,
                 settings=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("리포트 생성")
        self.setStyleSheet(STYLE_DLG)
        self.resize(420, 360)
        self.setWindowFlag(Qt.Window)

        self._result   = result
        self._gws      = gws
        self._nodes    = nodes
        self._heatmaps = heatmaps or []
        self._settings = settings or {}
        self._thread   = None
        self._worker   = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(10)

        # ── 현황 요약 ────────────────────────────────────────
        grp_info = QGroupBox("현재 분석 현황")
        grp_info.setStyleSheet(
            f"QGroupBox{{color:{MUTED};border:1px solid {BORDER};"
            f"border-radius:6px;margin-top:6px;padding-top:8px;}}"
            f"QGroupBox::title{{subcontrol-origin:margin;left:8px;}}")
        info_lay = QFormLayout(grp_info); info_lay.setSpacing(6)

        def _row(k, v, color=TEXT):
            lbl_k = QLabel(k)
            lbl_k.setStyleSheet(f"color:{MUTED};font-size:11px;")
            lbl_v = QLabel(v)
            lbl_v.setStyleSheet(f"color:{color};font-size:11px;font-weight:bold;")
            info_lay.addRow(lbl_k, lbl_v)

        if self._result:
            pct     = self._result.coverage_pct
            n_cov   = self._result.n_covered
            n_total = self._result.n_total
            n_gws   = len(self._gws)
            col     = '#00C94A' if pct >= 90 else '#FFD700' if pct >= 70 else '#FF4444'
            _row("전체 커버리지",  f"{pct:.1f}%", col)
            _row("커버 Node",      f"{n_cov} / {n_total} 개")
            _row("활성 GW",        f"{n_gws} 개")
        else:
            _row("커버리지 분석", "미실행", '#FF4444')

        lay.addWidget(grp_info)

        # ── 출력 옵션 ────────────────────────────────────────
        grp_opt = QGroupBox("출력 옵션")
        grp_opt.setStyleSheet(grp_info.styleSheet())
        opt_lay = QVBoxLayout(grp_opt); opt_lay.setSpacing(6)

        self.chk_excel = QCheckBox("Excel 내보내기 (.xlsx)")
        self.chk_pdf   = QCheckBox("PDF 리포트 생성 (.pdf)")
        for chk in [self.chk_excel, self.chk_pdf]:
            chk.setChecked(True)
            chk.setStyleSheet(
                f"QCheckBox{{color:{TEXT};font-size:11px;spacing:6px;}}"
                f"QCheckBox::indicator{{width:16px;height:16px;"
                f"border:1px solid {BORDER};border-radius:3px;"
                f"background:{PANEL};}}"
                f"QCheckBox::indicator:checked{{background:#1d4a1d;"
                f"border-color:#2a6a2a;}}")
            opt_lay.addWidget(chk)

        lay.addWidget(grp_opt)

        # ── 진행바 ───────────────────────────────────────────
        self.prog = QProgressBar()
        self.prog.setRange(0, 100); self.prog.setValue(0)
        self.prog.setFixedHeight(14)
        self.prog.setStyleSheet(
            f"QProgressBar{{background:{DARK};border:1px solid {BORDER};"
            f"border-radius:6px;text-align:center;color:{TEXT};font-size:10px;}}"
            f"QProgressBar::chunk{{background:#4f8ef7;border-radius:6px;}}")
        self.prog.setTextVisible(True)
        self.prog.setFormat("%p%")
        lay.addWidget(self.prog)

        self.lbl_status = QLabel("리포트를 생성할 항목을 선택 후 생성 버튼을 누르세요.")
        self.lbl_status.setStyleSheet(f"color:{MUTED};font-size:10px;")
        self.lbl_status.setWordWrap(True)
        lay.addWidget(self.lbl_status)

        # ── 버튼 ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self.btn_gen = QPushButton("📄 리포트 생성")
        self.btn_gen.setStyleSheet(
            f"QPushButton{{background:#1d4a1d;color:#7ae87a;"
            f"border:1px solid #2a6a2a;border-radius:5px;"
            f"padding:7px 20px;font-size:12px;font-weight:bold;}}"
            f"QPushButton:hover{{background:#256a25;}}"
            f"QPushButton:disabled{{background:#1a1a1a;color:{MUTED};}}")
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet(
            f"QPushButton{{background:{PANEL};color:{MUTED};"
            f"border:1px solid {BORDER};border-radius:5px;"
            f"padding:7px 16px;font-size:12px;}}"
            f"QPushButton:hover{{color:{TEXT};}}")
        self.btn_gen.clicked.connect(self._generate)
        btn_close.clicked.connect(self.close)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        btn_row.addWidget(self.btn_gen)
        lay.addLayout(btn_row)

        if not self._result:
            self.btn_gen.setEnabled(False)
            self.lbl_status.setText("커버리지 분석을 먼저 실행하세요.")

    def _generate(self):
        if not self.chk_excel.isChecked() and not self.chk_pdf.isChecked():
            QMessageBox.information(self, "알림", "출력 옵션을 하나 이상 선택하세요.")
            return

        if self.chk_excel.isChecked():
            path, _ = QFileDialog.getSaveFileName(
                self, "Excel 저장", "coverage_report.xlsx",
                "Excel (*.xlsx)")
            if path:
                self._run_worker('excel', path)

        if self.chk_pdf.isChecked():
            path, _ = QFileDialog.getSaveFileName(
                self, "PDF 저장", "coverage_report.pdf",
                "PDF (*.pdf)")
            if path:
                self._run_worker('pdf', path)

    def _run_worker(self, mode, path):
        self.btn_gen.setEnabled(False)
        self.prog.setValue(0)
        self.lbl_status.setText(f"{'Excel' if mode=='excel' else 'PDF'} 생성 중...")

        w = ReportWorker(mode, path, self._result, self._gws,
                         self._nodes, self._heatmaps, self._settings)
        t = QThread()
        self._thread = t
        self._worker = w
        w.moveToThread(t)
        t.started.connect(w.run)
        w.sig_progress.connect(self._on_progress)
        w.sig_done.connect(self._on_done)
        w.sig_err.connect(self._on_err)
        w.sig_done.connect(t.quit)
        w.sig_err.connect(t.quit)
        t.start()

    def _on_progress(self, pct, msg):
        self.prog.setValue(pct)
        self.lbl_status.setText(msg)

    def _on_done(self, path):
        self.prog.setValue(100)
        self.lbl_status.setText(f"완료: {os.path.basename(path)}")
        self.btn_gen.setEnabled(True)
        QMessageBox.information(
            self, "완료",
            f"리포트가 생성되었습니다.\n{path}")

    def _on_err(self, msg):
        self.prog.setValue(0)
        self.lbl_status.setText("오류 발생 — 콘솔 확인")
        self.btn_gen.setEnabled(True)
        print(f"[REPORT ERROR]\n{msg}")
        QMessageBox.critical(self, "오류",
            f"리포트 생성 실패.\n필요 라이브러리: openpyxl, reportlab\n\n{msg[:200]}")